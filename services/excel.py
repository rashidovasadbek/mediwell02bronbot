# -*- coding: utf-8 -*-
"""Excel yasash — hammasi xotirada, diskka hech narsa yozilmaydi.

farm_bot spesifikatsiyani ish papkasiga `Order_{apteka}.xlsx` nomi bilan
yozib, yuborgach o'chirardi. Ikki menejer bir aptekaga bir vaqtda
buyurtma qilsa fayl nomi to'qnashib, biri ikkinchisinikini yuborardi.
Bu yerda faqat bytes qaytariladi.

pandas ishlatilmaydi: farm_bot uni faqat DataFrame yasash uchun chaqirardi,
yozishni baribir openpyxl qilardi.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.pricing import Totals

HEADER_FILL = PatternFill("solid", start_color="002060", end_color="002060")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREY_FILL = PatternFill("solid", start_color="E9E9E9", end_color="E9E9E9")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

MONEY_FMT = "#,##0"
SPEC_COLUMNS = [
    ("№", 6),
    ("Mahsulot nomi", 45),
    ("O'lchov", 12),
    ("Miqdori", 12),
    ("Narxi (NDSsiz)", 18),
    ("NDS %", 10),
    ("NDS summasi", 18),
    ("Jami summa", 18),
]


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_spec_excel(totals: Totals, company, pharmacy) -> bytes:
    """Bron spesifikatsiyasi — sotuvchi/xaridor rekvizitlari bilan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bron"

    # --- Sarlavha ---
    for col, (title, width) in enumerate(SPEC_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        # farm_botda chr(64 + col) ishlatilardi — 26 ustundan oshsa buziladi
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- Qatorlar ---
    row = 2
    for i, ln in enumerate(totals.lines, 1):
        values = [i, ln.name, ln.unit, ln.quantity,
                  ln.price_no_nds, ln.nds_rate, ln.nds_sum, ln.line_total]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = LEFT if col == 2 else CENTER
            if col in (5, 7, 8):
                cell.number_format = MONEY_FMT
        row += 1

    # --- Jami ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    total_label = ws.cell(row=row, column=1, value="TO'LOV UCHUN JAMI:")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_cell = ws.cell(row=row, column=8, value=totals.grand_total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = MONEY_FMT
    total_cell.border = BORDER

    # --- Imzolar ---
    sign_row = row + 3
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=8)
    title = ws.cell(row=sign_row, column=1, value="TOMONLAR IMZOLARI")
    title.alignment = CENTER
    title.font = Font(bold=True, size=12)

    ws.merge_cells(start_row=sign_row + 1, start_column=1, end_row=sign_row + 1, end_column=4)
    ws.merge_cells(start_row=sign_row + 1, start_column=5, end_row=sign_row + 1, end_column=8)
    for col, label in ((1, "Sotuvchi:"), (5, "Xaridor:")):
        cell = ws.cell(row=sign_row + 1, column=col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.fill = GREY_FILL
        cell.border = BORDER

    seller = [
        company["name"],
        f"Manzil: {company['address'] or '—'}",
        f"H/r: {company['account_no'] or '—'}",
        f"Bank: {company['bank_name'] or '—'}",
        f"INN: {company['inn'] or '—'}   MFO: {company['mfo'] or '—'}",
        f"Direktor: {company['director'] or '—'}",
    ]
    contract_date = pharmacy["contract_date"]
    buyer = [
        pharmacy["name"],
        f"Region: {pharmacy['region_name']}",
        f"Menejer: {pharmacy['manager_name'] or '—'}",
        f"Shartnoma: №{pharmacy['contract_no']}"
        f"  ({contract_date.strftime('%d.%m.%Y') if contract_date else '—'})",
        f"INN: {pharmacy['inn']}",
        f"Tel: {pharmacy['phone'] or '—'}",
        "Direktor:",
        "M.P. (muhr uchun joy)",
    ]

    for i in range(max(len(seller), len(buyer))):
        r = sign_row + 2 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        for col, source in ((1, seller), (5, buyer)):
            cell = ws.cell(row=r, column=col, value=source[i] if i < len(source) else "")
            cell.alignment = LEFT
            cell.font = Font(size=10)
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER

    return _save(wb)


PHARMACY_COLUMNS = [
    ("№", 6),
    ("Apteka nomi", 42),
    ("Viloyat", 18),
    ("INN", 16),
    ("Telefon", 18),
    ("Shartnoma №", 18),
    ("Shartnoma sanasi", 18),
    ("Menejer", 24),
]


def build_pharmacies_excel(rows) -> bytes:
    """Aptekalar ro'yxati."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aptekalar"

    for col, (title, width) in enumerate(PHARMACY_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r["name"],
            r["region_name"],
            r["inn"],              # matn: boshidagi nol yo'qolmasin
            r["phone"] or "—",
            r["contract_no"] or "—",
            r["contract_date"] or "—",
            r["manager_name"] or "—",
        ])
        if r["contract_date"]:
            ws.cell(row=i + 1, column=7).number_format = "DD.MM.YYYY"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PHARMACY_COLUMNS))}{len(rows) + 1}"
    return _save(wb)
